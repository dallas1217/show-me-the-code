#-*- coding=utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook


with open('student_info.txt','r') as f:
    data = f.read()
    wb = Workbook()
    ws = wb.create_sheet(0)
    ws.title = 'student'

    for keys, values in eval(data).items():
        ws['A%s' % keys].value = keys
        for i in range(0, 4):
            ws['%s%s' % (chr(i+66), keys)].value = values[i]
            
    
    wb.save('0014.xlsx')
~                                
